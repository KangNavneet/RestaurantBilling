# import openpyxl module
import openpyxl
import random
from openpyxl  import *
# import Font function from openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from connection import *
from tkinter.messagebox import *
from addToCart import *





class excelFile:


    # def generateexcel(self):
        # Call a Workbook() function of openpyxl
        # to create a new blank Workbook object


    def generateexcel(self,x):


        self.mainList=x



        wb=openpyxl.Workbook()

        print(self.mainList)
        '''VARIABLE DECLARARTION'''
        '''ORDER DETAIL FROM DATABASE'''


        '''MENU DETAIL FROM MAINLIST'''
        menuItemListConst = []
        priceListConst=[]
        quantityListConst=[]
        totalPriceListConst=[]



        '''DATA DETAILS'''
        cur=con.cursor()
        query="Select * from bill "
        cur.execute(query)
        self.data1=cur.fetchall()
        for i in self.data1:
            billId=i[0]


        query="Select * from bill where billid='"+str(billId)+"'"
        cur.execute(query)
        self.billData=cur.fetchone()


        typeOfOrder = str(self.billData[1])
        modeOfPayment= str(self.billData[3])
        netAmount = str(self.billData[7])
        cashCollected = str(self.billData[4])
        cashReturned = str(self.billData[5])

        gst = str(self.billData[8])
        totalBill = float(netAmount)/(1+float(gst)*0.01)




        for i in range(0,len(self.mainList)):
          for j in range(0,len(self.mainList[i])):
            print("********MAINLIST LOOP!**********")
            print(self.mainList[i][j])

        print("******************")

        for i in range(1, len(self.mainList)):

            '''ORDER DETAIL'''
            menuItem = self.mainList[i][1]
            menuItemListConst.insert(i-1,menuItem)
            price = self.mainList[i][2]
            priceListConst.insert(i - 1, price)
            quantity = self.mainList[i][3]
            quantityListConst.insert(i - 1, quantity)
            totalPrice = self.mainList[i][4]
            totalPriceListConst.insert(i-1,totalPrice)

        '''DATA DETAILS ENDS HERE'''

        # Get workbook active sheet
        # from the active attribute.
        sheet = wb.active

        # set the height of the row
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[2].height = 25
        sheet.row_dimensions[3].height = 25
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[5].height = 25
        sheet.row_dimensions[6].height = 25
        sheet.row_dimensions[7].height = 25
        sheet.row_dimensions[8].height = 25
        sheet.row_dimensions[9].height = 25
        sheet.row_dimensions[10].height = 25
        sheet.row_dimensions[11].height = 25

        # set the width of the column
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 30
        sheet.column_dimensions['K'].width = 30
        index = 1
        '''MENU DETAIL'''
        sheet.cell(row=index, column=1).value = 'MENU NAME'
        sheet.cell(row=index, column=1).font = Font(size=18, bold=True)


        sheet.cell(row=index, column=2).value = 'PRICE OF MENU'
        sheet.cell(row=index, column=2).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=3).value = 'QUANTITY'
        sheet.cell(row=index, column=3).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=4).value = 'TOTAL PRICE'
        sheet.cell(row=index, column=4).font = Font(size=18, bold=True)

        print(menuItemListConst)
        print(priceListConst)
        print(quantityListConst)
        print(totalPriceListConst)

        '''DATA INSERTION IN EXCEL'''
        index = index + 1
        for i in range(0,len(menuItemListConst)):

            sheet.cell(row=index, column=1).value =str(menuItemListConst[i])
            sheet.cell(row=index, column=2).value =str(priceListConst[i])
            sheet.cell(row=index, column=3).value =str(quantityListConst[i])
            sheet.cell(row=index, column=4).value =str(totalPriceListConst[i])


            index=index+1


        index=index+1
        '''ORDER DETAIL'''
        index=index+1
        sheet.cell(row=index, column=1).value = 'TYPE OF ORDER'
        sheet.cell(row=index, column=1).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=2).value = 'MODE OF PAYMENT'
        sheet.cell(row=index, column=2).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=3).value = 'BILL'
        sheet.cell(row=index, column=3).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=4).value = 'TOTAL GST'
        sheet.cell(row=index, column=4).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=5).value = 'TOTAL BILL'
        sheet.cell(row=index, column=5).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=6).value = 'CASH COLLECTED'
        sheet.cell(row=index, column=6).font = Font(size=18, bold=True)

        sheet.cell(row=index, column=7).value = 'CASH RETURNED'
        sheet.cell(row=index, column=7).font = Font(size=18, bold=True)


        index = index + 1
        sheet.cell(row=index, column=1).value = str(typeOfOrder)
        sheet.cell(row=index, column=2).value = str(modeOfPayment)
        sheet.cell(row=index, column=3).value = str(totalBill)
        sheet.cell(row=index, column=4).value = str(gst)
        sheet.cell(row=index, column=5).value = str(netAmount)
        sheet.cell(row=index, column=6).value = str(cashCollected)
        sheet.cell(row=index, column=7).value = str(cashReturned)


        # save the file
        name = str(random.randint(0, 100))

        wb.save('bills//' + name + '.xlsx')
        showinfo('', 'excel file is stored at' + 'bills//' + name + '.xlsx')





    def __init__(self,mainList):
        self.generateexcel(mainList)


